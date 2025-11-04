#!/usr/bin/env python3
"""
Flask API Backend for Voter List OCR
Handles PDF uploads and processes them using voter_list_ocr.py
"""

import os
import sys
from pathlib import Path
from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from werkzeug.utils import secure_filename
from datetime import datetime
import threading
from voter_list_ocr import VoterListOCR
from openpyxl import load_workbook, Workbook
import io
import base64
from pdf2image import convert_from_path
from PIL import Image

app = Flask(__name__, static_folder='static', template_folder='templates')
CORS(app)

# Configuration
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
ALLOWED_EXTENSIONS = {'pdf'}
MAX_FILE_SIZE = 500 * 1024 * 1024  # 500MB max file size

# Create necessary directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Store processing status
processing_status = {}


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/')
def index():
    """Serve the main frontend page"""
    return render_template('index.html')


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'message': 'Voter List OCR API is running'
    })


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Handle PDF file upload (no processing yet)"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Only PDF files are allowed.'}), 400

    try:
        # Secure filename and save
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)

        file.save(filepath)

        # Initialize processing status (but don't start processing yet)
        job_id = f"{timestamp}_{Path(filename).stem}"
        processing_status[job_id] = {
            'status': 'uploaded',
            'progress': 0,
            'message': 'File uploaded successfully',
            'filename': filename,
            'pdf_path': filepath,
            'output_file': None,
            'error': None
        }

        return jsonify({
            'success': True,
            'job_id': job_id,
            'message': 'File uploaded successfully',
            'filename': filename
        }), 200

    except Exception as e:
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500


@app.route('/api/preview/<job_id>', methods=['GET'])
def preview_pdf(job_id):
    """Generate PDF thumbnails for preview"""
    if job_id not in processing_status:
        return jsonify({'error': 'Job ID not found'}), 404

    status = processing_status[job_id]

    if 'pdf_path' not in status or not os.path.exists(status['pdf_path']):
        return jsonify({'error': 'PDF file not found'}), 404

    try:
        pdf_path = status['pdf_path']

        # Convert PDF to images with low DPI for thumbnails
        images = convert_from_path(pdf_path, dpi=100)
        total_pages = len(images)

        # Generate thumbnails
        thumbnails = []
        for i, image in enumerate(images, 1):
            # Resize to thumbnail (200px height)
            aspect_ratio = image.width / image.height
            thumbnail_height = 200
            thumbnail_width = int(thumbnail_height * aspect_ratio)
            thumbnail = image.resize((thumbnail_width, thumbnail_height), Image.Resampling.LANCZOS)

            # Convert to base64
            buffer = io.BytesIO()
            thumbnail.save(buffer, format='PNG')
            buffer.seek(0)
            image_base64 = base64.b64encode(buffer.read()).decode('utf-8')

            thumbnails.append({
                'page': i,
                'data': image_base64
            })

        return jsonify({
            'success': True,
            'thumbnails': thumbnails,
            'total_pages': total_pages
        }), 200

    except Exception as e:
        return jsonify({'error': f'Failed to generate previews: {str(e)}'}), 500


@app.route('/api/process/<job_id>', methods=['POST'])
def process_pdf(job_id):
    """Start processing PDF with configuration"""
    if job_id not in processing_status:
        return jsonify({'error': 'Job ID not found'}), 404

    status = processing_status[job_id]

    if 'pdf_path' not in status or not os.path.exists(status['pdf_path']):
        return jsonify({'error': 'PDF file not found'}), 404

    try:
        # Get configuration from request
        config = request.get_json()
        matadaar_sangh = config.get('matadaar_sangh', '')
        election_type = config.get('election_type', '')
        ward_number = config.get('ward_number', '')
        start_page = config.get('start_page', 1)
        end_page = config.get('end_page', None)
        language = config.get('language', 'marathi')

        # Store configuration in processing status
        processing_status[job_id]['config'] = config
        processing_status[job_id]['status'] = 'processing'
        processing_status[job_id]['progress'] = 10
        processing_status[job_id]['message'] = 'Starting processing with configuration...'

        # Start processing in background thread
        thread = threading.Thread(
            target=process_pdf_background,
            args=(status['pdf_path'], job_id, status['filename'], config)
        )
        thread.daemon = True
        thread.start()

        return jsonify({
            'success': True,
            'message': 'Processing started'
        }), 200

    except Exception as e:
        return jsonify({'error': f'Failed to start processing: {str(e)}'}), 500


def process_pdf_background(pdf_path, job_id, original_filename, config=None):
    """Process PDF in background thread"""
    try:
        # Extract configuration
        if config is None:
            config = {}

        matadaar_sangh = config.get('matadaar_sangh', '')
        election_type = config.get('election_type', '')
        ward_number = config.get('ward_number', '')
        start_page = config.get('start_page', 1)
        end_page = config.get('end_page', None)

        # Update status
        processing_status[job_id]['status'] = 'processing'
        processing_status[job_id]['message'] = 'Converting PDF to images...'
        processing_status[job_id]['progress'] = 10

        # Generate output filename
        output_filename = f"{Path(original_filename).stem}_voters.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{job_id}_{output_filename}")

        # Initialize OCR processor with configuration
        ocr = VoterListOCR(
            pdf_path,
            output_path,
            matadaar_sangh=matadaar_sangh,
            election_type=election_type,
            ward_number=ward_number
        )

        # Update status
        processing_status[job_id]['message'] = 'Initializing Excel file...'
        processing_status[job_id]['progress'] = 20

        # Initialize Excel file
        ocr.initialize_excel_file()

        # Update status
        processing_status[job_id]['message'] = 'Extracting text from PDF pages...'
        processing_status[job_id]['progress'] = 30

        # Extract text from specified page range
        page_texts = ocr.extract_text_from_pdf(start_page=start_page, end_page=end_page)

        total_pages = len(page_texts)
        processing_status[job_id]['message'] = f'Processing {total_pages} pages (from page {start_page} to {end_page or "end"})...'
        processing_status[job_id]['progress'] = 50

        # Process each page
        all_voters_count = 0
        for i, text in enumerate(page_texts, 1):
            try:
                # Parse voters from this page
                voters = ocr.parse_voter_info(text)

                if voters:
                    # Save to Excel incrementally
                    ocr.append_voters_to_excel(voters)
                    all_voters_count += len(voters)

                # Update progress
                progress = 50 + int((i / total_pages) * 40)
                processing_status[job_id]['progress'] = progress
                processing_status[job_id]['message'] = f'Processed page {i}/{total_pages} - Found {all_voters_count} voters'

            except Exception as e:
                print(f"Error processing page {i}: {e}")
                continue

        # Finalize
        processing_status[job_id]['status'] = 'completed'
        processing_status[job_id]['progress'] = 100
        processing_status[job_id]['message'] = f'Processing completed! Extracted {all_voters_count} voter records.'
        processing_status[job_id]['output_file'] = output_filename
        processing_status[job_id]['output_path'] = output_path
        processing_status[job_id]['total_records'] = all_voters_count

        # Clean up uploaded file after processing
        try:
            os.remove(pdf_path)
        except:
            pass

    except Exception as e:
        processing_status[job_id]['status'] = 'error'
        processing_status[job_id]['error'] = str(e)
        processing_status[job_id]['message'] = f'Error: {str(e)}'
        print(f"Processing error: {e}")


@app.route('/api/status/<job_id>', methods=['GET'])
def get_status(job_id):
    """Get processing status for a job"""
    if job_id not in processing_status:
        return jsonify({'error': 'Job ID not found'}), 404
    
    status = processing_status[job_id]
    return jsonify(status), 200


@app.route('/api/data/<job_id>', methods=['GET'])
def get_data(job_id):
    """Get the processed Excel data as JSON"""
    if job_id not in processing_status:
        return jsonify({'error': 'Job ID not found'}), 404

    status = processing_status[job_id]

    if status['status'] != 'completed' or 'output_path' not in status:
        return jsonify({'error': 'Data not ready'}), 400

    if not os.path.exists(status['output_path']):
        return jsonify({'error': 'Output file not found'}), 404

    try:
        # Load the Excel file
        workbook = load_workbook(status['output_path'])
        sheet = workbook.active

        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)

        # Get all data rows
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Convert row to dictionary
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    # Convert None to empty string for JSON serialization
                    row_dict[headers[i]] = value if value is not None else ''
            data.append(row_dict)

        workbook.close()

        return jsonify({
            'success': True,
            'headers': headers,
            'data': data,
            'total_records': len(data)
        }), 200

    except Exception as e:
        return jsonify({'error': f'Failed to read data: {str(e)}'}), 500


@app.route('/api/download/<job_id>', methods=['GET'])
def download_file(job_id):
    """Download the processed Excel file"""
    if job_id not in processing_status:
        return jsonify({'error': 'Job ID not found'}), 404

    status = processing_status[job_id]

    if status['status'] != 'completed' or 'output_path' not in status:
        return jsonify({'error': 'File not ready for download'}), 400

    if not os.path.exists(status['output_path']):
        return jsonify({'error': 'Output file not found'}), 404

    try:
        return send_file(
            status['output_path'],
            as_attachment=True,
            download_name=status['output_file'],
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': f'Download failed: {str(e)}'}), 500


@app.route('/api/download-filtered/<job_id>', methods=['POST'])
def download_filtered(job_id):
    """Generate and download Excel file from filtered data"""
    if job_id not in processing_status:
        return jsonify({'error': 'Job ID not found'}), 404

    status = processing_status[job_id]

    if status['status'] != 'completed':
        return jsonify({'error': 'Processing not completed'}), 400

    try:
        # Get filtered data from request
        data = request.get_json()
        headers = data.get('headers', [])
        rows = data.get('data', [])

        if not headers or not rows:
            return jsonify({'error': 'No data provided'}), 400

        # Create new workbook
        workbook = Workbook()
        sheet = workbook.active

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))

        # Save to BytesIO
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        workbook.close()

        # Generate filename
        original_filename = status.get('filename', 'voters')
        filename = f"{Path(original_filename).stem}_filtered.xlsx"

        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': f'Download failed: {str(e)}'}), 500


@app.route('/api/list-jobs', methods=['GET'])
def list_jobs():
    """List all processing jobs"""
    return jsonify({
        'jobs': list(processing_status.keys()),
        'count': len(processing_status)
    }), 200


if __name__ == '__main__':
    print("="*70)
    print("Voter List OCR Web Application")
    print("="*70)
    print(f"Upload folder: {app.config['UPLOAD_FOLDER']}")
    print(f"Output folder: {app.config['OUTPUT_FOLDER']}")
    print(f"Max file size: {MAX_FILE_SIZE / (1024*1024):.0f} MB")
    print("="*70)
    print("\nStarting server...")
    print("\n" + "="*70)
    print("SERVER IS RUNNING!")
    print("="*70)
    print("\n🌐 Open your browser and go to:")
    print("   http://localhost:8080")
    print("   OR")
    print("   http://127.0.0.1:8080")
    print("\n💡 If the page doesn't load:")
    print("   1. Check your browser's developer console (F12)")
    print("   2. Try a different browser")
    print("   3. Clear browser cache (Cmd+Shift+R or Ctrl+Shift+R)")
    print("   4. Try incognito/private mode")
    print("\n" + "="*70)
    print("Press Ctrl+C to stop the server")
    print("="*70 + "\n")
    
    try:
        app.run(host='0.0.0.0', port=8080, debug=True)
    except OSError as e:
        if "Address already in use" in str(e):
            print(f"\n❌ ERROR: Port 8080 is already in use!")
            print("   Solution: Kill the process using port 8080:")
            print("   lsof -ti:8080 | xargs kill -9")
            print("   Then run the app again.\n")
        else:
            print(f"\n❌ ERROR: {e}\n")
        sys.exit(1)

