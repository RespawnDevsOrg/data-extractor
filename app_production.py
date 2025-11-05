
#!/usr/bin/env python3
"""
Production Flask API Backend for Voter List OCR
Optimized for production deployment with enhanced security and performance
"""

import os
import sys
import logging
from pathlib import Path
from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from werkzeug.utils import secure_filename
from datetime import datetime
import threading
from voter_list_ocr import VoterListOCR
from openpyxl import load_workbook, Workbook
import io
import base64
from pdf2image import convert_from_path
from PIL import Image

# Production logging configuration
LOG_FILE = 'voter_ocr_app.log'

# Configure production logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.FileHandler(LOG_FILE, mode='a', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger(__name__)
logger.info("="*70)
logger.info("VOTER LIST OCR APPLICATION STARTED (PRODUCTION MODE)")
logger.info(f"Log file: {os.path.abspath(LOG_FILE)}")
logger.info("="*70)

app = Flask(__name__, static_folder='static', template_folder='templates')

# Production CORS configuration
CORS(app, origins=["http://localhost:8080", "http://127.0.0.1:8080"])

# Production Configuration
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
ALLOWED_EXTENSIONS = {'pdf'}
MAX_FILE_SIZE = 500 * 1024 * 1024  # 500MB max file size

# Create necessary directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

logger.info(f"Upload folder: {os.path.abspath(UPLOAD_FOLDER)}")
logger.info(f"Output folder: {os.path.abspath(OUTPUT_FOLDER)}")

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Production security headers
@app.after_request
def after_request(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    return response

# Store processing status
processing_status = {}

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    """Serve the main frontend page"""
    return render_template('index.html')

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'message': 'Voter List OCR API is running (Production Mode)',
        'version': '1.0.0'
    })

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Handle PDF file upload (no processing yet)"""
    if 'file' not in request.files:
        logger.warning("Upload attempt without file")
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']

    if file.filename == '':
        logger.warning("Upload attempt with empty filename")
        return jsonify({'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        logger.warning(f"Upload attempt with invalid file type: {file.filename}")
        return jsonify({'error': 'Invalid file type. Only PDF files are allowed.'}), 400

    try:
        # Secure filename and save
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)

        file.save(filepath)
        logger.info(f"File uploaded successfully: {unique_filename}")

        # Initialize processing status
        job_id = f"{timestamp}_{Path(filename).stem}"
        processing_status[job_id] = {
            'status': 'uploaded',
            'progress': 0,
            'message': 'File uploaded successfully',
            'filename': filename,
            'pdf_path': filepath,
            'output_file': None,
            'error': None
        }

        return jsonify({
            'success': True,
            'job_id': job_id,
            'message': 'File uploaded successfully',
            'filename': filename
        }), 200

    except Exception as e:
        logger.error(f"Upload failed: {str(e)}")
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500

@app.route('/api/preview/<job_id>', methods=['GET'])
def preview_pdf(job_id):
    """Generate PDF thumbnails for preview"""
    if job_id not in processing_status:
        logger.warning(f"Preview request for unknown job_id: {job_id}")
        return jsonify({'error': 'Job ID not found'}), 404

    status = processing_status[job_id]

    if 'pdf_path' not in status or not os.path.exists(status['pdf_path']):
        logger.error(f"PDF file not found for job_id: {job_id}")
        return jsonify({'error': 'PDF file not found'}), 404

    try:
        pdf_path = status['pdf_path']
        logger.info(f"Generating preview for: {pdf_path}")

        # Convert PDF to images with low DPI for thumbnails
        images = convert_from_path(pdf_path, dpi=100)
        total_pages = len(images)

        # Generate thumbnails
        thumbnails = []
        for i, image in enumerate(images, 1):
            # Resize to thumbnail (200px height)
            aspect_ratio = image.width / image.height
            thumbnail_height = 200
            thumbnail_width = int(thumbnail_height * aspect_ratio)
            thumbnail = image.resize((thumbnail_width, thumbnail_height), Image.Resampling.LANCZOS)

            # Convert to base64
            buffer = io.BytesIO()
            thumbnail.save(buffer, format='PNG')
            buffer.seek(0)
            image_base64 = base64.b64encode(buffer.read()).decode('utf-8')

            thumbnails.append({
                'page': i,
                'data': image_base64
            })

        logger.info(f"Generated {len(thumbnails)} thumbnails for job_id: {job_id}")
        return jsonify({
            'success': True,
            'thumbnails': thumbnails,
            'total_pages': total_pages
        }), 200

    except Exception as e:
        logger.error(f"Failed to generate previews for job_id {job_id}: {str(e)}")
        return jsonify({'error': f'Failed to generate previews: {str(e)}'}), 500

@app.route('/api/process/<job_id>', methods=['POST'])
def process_pdf(job_id):
    """Start processing PDF with configuration"""
    if job_id not in processing_status:
        logger.warning(f"Process request for unknown job_id: {job_id}")
        return jsonify({'error': 'Job ID not found'}), 404

    status = processing_status[job_id]

    if 'pdf_path' not in status or not os.path.exists(status['pdf_path']):
        logger.error(f"PDF file not found for processing job_id: {job_id}")
        return jsonify({'error': 'PDF file not found'}), 404

    try:
        # Get configuration from request
        config = request.get_json()
        matadaar_sangh = config.get('matadaar_sangh', '')
        election_type = config.get('election_type', '')
        ward_number = config.get('ward_number', '')
        start_page = config.get('start_page', 1)
        end_page = config.get('end_page', None)
        language = config.get('language', 'marathi')

        logger.info(f"Starting processing for job_id: {job_id} with config: {config}")

        # Store configuration in processing status
        processing_status[job_id]['config'] = config
        processing_status[job_id]['status'] = 'processing'
        processing_status[job_id]['progress'] = 10
        processing_status[job_id]['message'] = 'Starting processing with configuration...'

        # Start processing in background thread
        thread = threading.Thread(
            target=process_pdf_background,
            args=(status['pdf_path'], job_id, status['filename'], config)
        )
        thread.daemon = True
        thread.start()

        return jsonify({
            'success': True,
            'message': 'Processing started'
        }), 200

    except Exception as e:
        logger.error(f"Failed to start processing for job_id {job_id}: {str(e)}")
        return jsonify({'error': f'Failed to start processing: {str(e)}'}), 500

def process_pdf_background(pdf_path, job_id, original_filename, config=None):
    """Process PDF in background thread with production optimizations"""
    import gc
    import time

    try:
        logger.info(f"Background processing started for job_id: {job_id}")
        
        # Extract configuration
        if config is None:
            config = {}

        matadaar_sangh = config.get('matadaar_sangh', '')
        election_type = config.get('election_type', '')
        ward_number = config.get('ward_number', '')
        start_page = config.get('start_page', 1)
        end_page = config.get('end_page', None)

        # Track start time for elapsed time calculation
        start_time = time.time()
        processing_status[job_id]['start_time'] = start_time
        processing_status[job_id]['elapsed_time'] = 0

        # Update status
        processing_status[job_id]['status'] = 'processing'
        processing_status[job_id]['message'] = 'Initializing OCR processor...'
        processing_status[job_id]['progress'] = 5

        # Generate output filename
        output_filename = f"{Path(original_filename).stem}_voters.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{job_id}_{output_filename}")

        # Initialize OCR processor
        ocr = VoterListOCR(
            pdf_path,
            output_path,
            matadaar_sangh=matadaar_sangh,
            election_type=election_type,
            ward_number=ward_number
        )

        # Update status
        processing_status[job_id]['message'] = 'Initializing Excel file...'
        processing_status[job_id]['progress'] = 10

        # Initialize Excel file
        ocr.initialize_excel_file()

        # Update status
        processing_status[job_id]['message'] = 'Starting OCR processing...'
        processing_status[job_id]['progress'] = 15

        # Process PDF incrementally
        try:
            page_range_msg = f" (Pages: {start_page or 1} to {end_page or 'end'})" if (start_page or end_page) else ""
            processing_status[job_id]['message'] = f'Processing PDF{page_range_msg}...'
            processing_status[job_id]['progress'] = 20

            all_voters_count = ocr.process_pdf_incrementally(start_page=start_page, end_page=end_page)
            
            # Update progress to completion
            processing_status[job_id]['progress'] = 90
            processing_status[job_id]['message'] = f'Processing completed - Found {all_voters_count} voters total'
            
        except Exception as e:
            logger.error(f"Error in OCR processing for job_id {job_id}: {e}")
            raise e

        # Calculate total processing time
        end_time = time.time()
        total_time_seconds = end_time - start_time
        processing_status[job_id]['total_time'] = total_time_seconds

        # Calculate accuracy
        accuracy = 100.0
        if hasattr(ocr, 'total_patterns_found') and ocr.total_patterns_found > 0:
            id_success_rate = (ocr.total_valid_ids / ocr.total_patterns_found) * 100
            if ocr.total_records_saved > 0 and all_voters_count > 0:
                save_success_rate = (ocr.total_records_saved / all_voters_count) * 100
                accuracy = (id_success_rate + save_success_rate) / 2
            else:
                accuracy = id_success_rate

        # Finalize
        processing_status[job_id]['status'] = 'completed'
        processing_status[job_id]['progress'] = 100
        processing_status[job_id]['message'] = f'Processing completed! Extracted {all_voters_count} voter records in {int(total_time_seconds)}s (Accuracy: {accuracy:.1f}%)'
        processing_status[job_id]['output_file'] = output_filename
        processing_status[job_id]['output_path'] = output_path
        processing_status[job_id]['total_records'] = all_voters_count
        processing_status[job_id]['accuracy'] = round(accuracy, 1)

        logger.info(f"Processing completed successfully for job_id: {job_id}, records: {all_voters_count}")

        # Clean up uploaded file after processing
        try:
            os.remove(pdf_path)
            logger.info(f"Cleaned up uploaded file: {pdf_path}")
        except:
            pass
        
        # Force final cleanup
        gc.collect()

    except MemoryError as e:
        logger.error(f"Memory error for job_id {job_id}: {e}")
        processing_status[job_id]['status'] = 'error'
        processing_status[job_id]['error'] = f'Memory error: {str(e)}. Try processing smaller page ranges.'
        processing_status[job_id]['message'] = f'Memory Error: {str(e)}'
        gc.collect()
        
    except Exception as e:
        logger.error(f"Processing error for job_id {job_id}: {e}")
        processing_status[job_id]['status'] = 'error'
        processing_status[job_id]['error'] = str(e)
        processing_status[job_id]['message'] = f'Error: {str(e)}'
        gc.collect()

@app.route('/api/status/<job_id>', methods=['GET'])
def get_status(job_id):
    """Get processing status for a job"""
    import time

    if job_id not in processing_status:
        return jsonify({'error': 'Job ID not found'}), 404

    status = processing_status[job_id].copy()

    # Calculate elapsed time if processing
    if status.get('status') == 'processing' and 'start_time' in status:
        elapsed = time.time() - status['start_time']
        status['elapsed_time'] = int(elapsed)
    elif status.get('status') == 'completed' and 'total_time' in status:
        status['elapsed_time'] = int(status['total_time'])

    return jsonify(status), 200

@app.route('/api/data/<job_id>', methods=['GET'])
def get_data(job_id):
    """Get the processed Excel data as JSON"""
    if job_id not in processing_status:
        return jsonify({'error': 'Job ID not found'}), 404

    status = processing_status[job_id]

    if status['status'] != 'completed' or 'output_path' not in status:
        return jsonify({'error': 'Data not ready'}), 400

    if not os.path.exists(status['output_path']):
        return jsonify({'error': 'Output file not found'}), 404

    try:
        # Load the Excel file
        workbook = load_workbook(status['output_path'])
        sheet = workbook.active

        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)

        # Get all data rows
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Convert row to dictionary
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    # Convert None to empty string for JSON serialization
                    row_dict[headers[i]] = value if value is not None else ''
            data.append(row_dict)

        workbook.close()

        return jsonify({
            'success': True,
            'headers': headers,
            'data': data,
            'total_records': len(data)
        }), 200

    except Exception as e:
        logger.error(f"Failed to read data for job_id {job_id}: {str(e)}")
        return jsonify({'error': f'Failed to read data: {str(e)}'}), 500

@app.route('/api/download/<job_id>', methods=['GET'])
def download_file(job_id):
    """Download the processed Excel file"""
    if job_id not in processing_status:
        return jsonify({'error': 'Job ID not found'}), 404

    status = processing_status[job_id]

    if status['status'] != 'completed' or 'output_path' not in status:
        return jsonify({'error': 'File not ready for download'}), 400

    if not os.path.exists(status['output_path']):
        return jsonify({'error': 'Output file not found'}), 404

    try:
        return send_file(
            status['output_path'],
            as_attachment=True,
            download_name=status['output_file'],
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Download failed for job_id {job_id}: {str(e)}")
        return jsonify({'error': f'Download failed: {str(e)}'}), 500

@app.route('/api/download-filtered/<job_id>', methods=['POST'])
def download_filtered(job_id):
    """Generate and download Excel file from filtered data"""
    if job_id not in processing_status:
        return jsonify({'error': 'Job ID not found'}), 404

    status = processing_status[job_id]

    if status['status'] != 'completed':
        return jsonify({'error': 'Processing not completed'}), 400

    try:
        # Get filtered data from request
        data = request.get_json()
        headers = data.get('headers', [])
        rows = data.get('data', [])

        if not headers or not rows:
            return jsonify({'error': 'No data provided'}), 400

        # Create new workbook
        workbook = Workbook()
        sheet = workbook.active

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))

        # Save to BytesIO
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        workbook.close()

        # Generate filename
        original_filename = status.get('filename', 'voters')
        filename = f"{Path(original_filename).stem}_filtered.xlsx"

        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Filtered download failed for job_id {job_id}: {str(e)}")
        return jsonify({'error': f'Download failed: {str(e)}'}), 500

@app.route('/api/list-jobs', methods=['GET'])
def list_jobs():
    """List all processing jobs"""
    return jsonify({
        'jobs': list(processing_status.keys()),
        'count': len(processing_status)
    }), 200

if __name__ == '__main__':
    print("="*70)
    print("Voter List OCR Web Application (PRODUCTION MODE)")
    print("="*70)
    print(f"Upload folder: {app.config['UPLOAD_FOLDER']}")
    print(f"Output folder: {app.config['OUTPUT_FOLDER']}")
    print(f"Max file size: {MAX_FILE_SIZE / (1024*1024):.0f} MB")
    print("\n🚀 PRODUCTION FEATURES:")
    print("   ✓ Enhanced security headers")
    print("   ✓ Production logging")
    print("   ✓ Error handling and monitoring")
    print("   ✓ Memory optimization")
    print("   ✓ File cleanup")
    print("="*70)
    print("\nStarting production server...")
    print("\n" + "="*70)
    print("SERVER IS RUNNING IN PRODUCTION MODE!")
    print("="*70)
    print("\n🌐 Access the application at:")
    print("   http://localhost:8080")
    print("   OR")
    print("   http://127.0.0.1:8080")
    print("\n" + "="*70)
    print("Press Ctrl+C to stop the server")
    print("="*70 + "\n")
    
    try:
        app.run(host='0.0.0.0', port=8080, debug=False)
    except OSError as e:
        if "Address already in use" in str(e):
            print(f"\n❌ ERROR: Port 8080 is already in use!")
            print("   Solution: Kill the process using port 8080:")
            print("   lsof -ti:8080 | xargs kill -9")
            print("   Then run the app again.\n")
        else:
            print(f"\n❌ ERROR: {e}\n")
        sys.exit(1)